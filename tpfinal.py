import os
import platform
from openpyxl import Workbook,load_workbook
import time

i=0

dest_archivo = "registro.xlsx"
if os.path.isfile(dest_archivo):
    libro = load_workbook("registro.xlsx")
else:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "registro"
    hoja.append(["cliente", "fecha", "combo S","combo D","combo T","flurby","total"])
    libro.save("registro.xlsx")



hoja = libro.active
hoja.title = "registro"



def verificar(dato):
    while dato == "":
        print("Error, valor vacio!")
        dato = input("Ingrese nuevamente: ")
    return dato
 


def convertir(valor):
    while valor.isdecimal() == False:
        print("Error, solo numeros enteros")
        valor = input("Ingrese nuevamente: ")
    valor = int(valor)
    return valor



def vuelto_total(abona,total):    
    while abona<total: 
        print("error en la suma, vuelva a ingresarla")
        abona = input(">>>")
        abona = verificar(abona)
        abona = convertir(abona)
    else:   
        total_total = abona - total
        print("vuelto $", total_total)
       


def menu(): 
 print(

"""

McDowell´s
Recuerda que siempre hay que recibir al cliente con una sonrisa :)
        
1 – Ingreso de nuevo pedido
2 – Cambio de turno
3 – Apagar sistema        
         
""")



def borrar_pantalla():
    sistema = platform.system()
    if sistema == "Windows":
        os.system("cls")
    else:
        os.system("clear")



def confirmacion(i,x): 
    print("¿Desea confirmar el pedido?")
    print("ingrese SI/NO ")
    sino = input(">>>")
    if sino == "S" or sino == "SI" or sino == "Si" or sino == "si" or sino == "sI" or sino == "s":
        print("PEDIDO CONFIRMADO")
        hoja.append([cliente, time.asctime(), s,d,t,f,total_total])
        libro.save("registro.xlsx")
        i = i + x
        print(i)
        return i
    elif sino == "N" or sino == "NO" or sino == "No" or sino == "no" or sino == "nO" or sino == "n":
        print("PEDIDO CANCELADO")
        i = i 
        return i
    else:   
        print("error en la opcion")
        confirmacion(i,x)
        return i
    
    

#######################################################################################################################################################################################

while True: 

    dest_archivo2 = open("registro.txt","a")
    fecha_ingreso = time.asctime()
    fecha_ingreso = str(fecha_ingreso)
   
    print("Bienvenido a McDowell´s ")
    encargado = input("Ingrese su nombre encargad@: ")
    encargado = verificar(encargado)
   
    log_ingreso = "IN  " + fecha_ingreso + " encargado " + encargado + "\n" 
    
    dest_archivo2.write(log_ingreso)
    dest_archivo2.close
    
    menu()
         
    opcion = input(">>> ")
    opcion = verificar(opcion)
    opcion = convertir(opcion)

    while opcion == 1:  

        s = input("Ingrese cantidad Combo S: ")
        s = verificar(s)
        s = convertir(s)
        d = input("Ingrese cantidad Combo D: ")
        d = verificar(d)
        d = convertir(d)
        t = input("Ingrese cantidad Combo T: ")
        t = verificar(t)
        t = convertir(t)
        f = input("Ingrese cantidad Flurby : ")
        f = verificar(f)
        f = convertir(f)

        cliente = input("ingrese el nombre del cliente: ")

        combo_simple   = 650
        combo_doble    = 700
        combo_triple   = 800
        postre_flurby  = 250

        total_simple = s * combo_simple
        total_doble  = d * combo_doble
        total_triple = t * combo_triple
        total_flurby = f * postre_flurby    
           
        total_total = total_simple + total_doble + total_triple + total_flurby
        print("total $:", total_total)                                            
       
        abona = input("Abona con $ >>> ")
        abona = verificar(abona)
        abona = convertir(abona)

        vuelto = vuelto_total(abona,total_total)
        
        
        i = confirmacion(i ,total_total)  

        
        menu()
        opcion = input(">>> ")
        opcion = verificar(opcion)
        opcion = convertir(opcion)

       
        borrar_pantalla()

        
    if opcion == 2:

        i = str(i)
        dest_archivo2 = open("registro.txt","a")
        fecha_salida =  time.asctime()
        fecha_salida =  str(fecha_salida)
        log_salida = "OUT " + fecha_salida + " encargado " + encargado + " $ " + i + "\n" 
        dest_archivo2.write(log_salida)
        dest_archivo2.close
        
        separador = "##############################################################################################" + "\n" 
        
        dest_archivo2.write(separador)
        dest_archivo2.close

        i = 0
        i = int(i)

        menu()
        borrar_pantalla()
        pass

    elif opcion == 3:   
        
        i = str(i)
        dest_archivo2 = open("registro.txt","a")
        fecha_salida =  time.asctime()
        fecha_salida =  str(fecha_salida)
        log_salida = fecha_salida + " encargado " + encargado  + " $ " + i +"\n" 
        dest_archivo2.write(log_salida)
        
        separador = "##############################################################################################" + "\n"
        dest_archivo2.write(separador)
        dest_archivo2.close
        print("adios", encargado)

        i = 0
        i = int(i)

        break
    
    else:   
        print("error en la opcion, ingresela nuevamente")
        opcion = input(">>> ")
        opcion = verificar(opcion)
        opcion = convertir(opcion)
        


        
#ARIEL BARTOLI y SEBASTIAN CHARRAS 

        
        


      




        
        


      


